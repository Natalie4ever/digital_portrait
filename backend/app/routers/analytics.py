# Step 4: 团队能力分析 API
# 5 个 API：overview / certificates / risks / emergency / export
from __future__ import annotations

import io
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, case
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models import (
    User,
    Profile,
    ContactInfo,
    ProfileSkillTag,
    QualificationInfo,
)
from app.schemas import (
    SkillCountItem,
    SkillDistributionResponse,
    GroupDensityItem,
    OverviewResponse,
    CertificateStatItem,
    CertificateStatResponse,
    RiskWarningItem,
    RisksResponse,
    EmergencyStatItem,
    EmergencyStatsResponse,
)
from app.auth import get_current_user, leader_effective_group
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter(prefix="/api/admin/analytics", tags=["团队能力分析"])

# 风险阈值：用户已确认
SKILL_RED_MAX = 1      # 仅 1 人 = 红
SKILL_YELLOW_MAX = 3   # 2-3 人 = 黄，4+ 人 = 绿
EMERGENCY_RED_MAX = 3  # 应急先锋队 < 3 = 红
EMERGENCY_YELLOW_MAX = 5  # < 5 = 黄，>= 5 = 绿


# ======================== 工具：组范围过滤 ========================

async def _scope_filter(q, db: AsyncSession, current_user: User):
    """组长自动限定本组"""
    if current_user.role == "leader":
        lg = leader_effective_group(current_user)
        if not lg:
            raise HTTPException(status_code=403, detail="组长未配置有效组别")
        q = q.where(User.group_name == lg)
    return q


# ======================== 1. overview ========================

@router.post("/overview", response_model=OverviewResponse)
async def overview(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ("admin", "leader"):
        raise HTTPException(status_code=403, detail="无权限")

    # 1) 技能分布
    skill_q = (
        select(ProfileSkillTag.tag_name, User.ehr_no)
        .join(Profile, ProfileSkillTag.profile_id == Profile.id)
        .join(User, Profile.user_id == User.id)
        .where(User.deleted_at.is_(None))
    )
    skill_q = await _scope_filter(skill_q, db, current_user)
    skill_rows = (await db.execute(skill_q)).all()
    skill_map: dict[str, list[str]] = {}
    for tag, ehr in skill_rows:
        skill_map.setdefault(tag, []).append(ehr)
    skills = [
        SkillCountItem(skill_name=tag, count=len(ehrs), holders=ehrs)
        for tag, ehrs in skill_map.items()
    ]
    skills.sort(key=lambda x: (-x.count, x.skill_name))

    # 2) 组别人数 + 应急先锋队人数
    grp_q = (
        select(User.group_name, func.count(User.id))
        .where(User.deleted_at.is_(None))
        .group_by(User.group_name)
    )
    grp_q = await _scope_filter(grp_q, db, current_user)
    grp_total = {g or "未分组": c for g, c in (await db.execute(grp_q)).all()}

    emg_q = (
        select(User.group_name, func.count(User.id))
        .join(Profile, Profile.user_id == User.id)
        .where(User.deleted_at.is_(None), Profile.is_emergency_staff == True)
        .group_by(User.group_name)
    )
    emg_q = await _scope_filter(emg_q, db, current_user)
    emg_map = {g or "未分组": c for g, c in (await db.execute(emg_q)).all()}

    groups = [
        GroupDensityItem(
            group_name=g,
            count=grp_total.get(g, 0),
            is_emergency_count=emg_map.get(g, 0),
        )
        for g in sorted(grp_total.keys())
    ]

    # 3) 总人数
    total_q = select(func.count(User.id)).where(User.deleted_at.is_(None))
    total_q = await _scope_filter(total_q, db, current_user)
    total_employees = (await db.execute(total_q)).scalar() or 0

    return OverviewResponse(
        skills=skills, groups=groups, total_employees=total_employees,
    )


# ======================== 2. certificates ========================

@router.post("/certificates", response_model=CertificateStatResponse)
async def certificates(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ("admin", "leader"):
        raise HTTPException(status_code=403, detail="无权限")

    q = (
        select(QualificationInfo.qualification_name, User.ehr_no)
        .join(Profile, QualificationInfo.profile_id == Profile.id)
        .join(User, Profile.user_id == User.id)
        .where(User.deleted_at.is_(None))
    )
    q = await _scope_filter(q, db, current_user)
    rows = (await db.execute(q)).all()
    cert_map: dict[str, list[str]] = {}
    for cert_name, ehr in rows:
        if not cert_name:
            continue
        cert_map.setdefault(cert_name, []).append(ehr)
    items = [
        CertificateStatItem(cert_name=c, count=len(ehrs), holders=ehrs)
        for c, ehrs in cert_map.items()
    ]
    items.sort(key=lambda x: (-x.count, x.cert_name))
    return CertificateStatResponse(
        items=items,
        total_certs=sum(len(e) for e in cert_map.values()),
    )


# ======================== 3. risks ========================

async def _calc_skill_risks(db: AsyncSession, current_user: User) -> list[RiskWarningItem]:
    """关键技能风险"""
    q = (
        select(ProfileSkillTag.tag_name, User.ehr_no, User.name)
        .join(Profile, ProfileSkillTag.profile_id == Profile.id)
        .join(User, Profile.user_id == User.id)
        .where(User.deleted_at.is_(None))
    )
    q = await _scope_filter(q, db, current_user)
    rows = (await db.execute(q)).all()
    skill_map: dict[str, list[dict]] = {}
    for tag, ehr, name in rows:
        skill_map.setdefault(tag, []).append({"ehr_no": ehr, "name": name})

    risks: list[RiskWarningItem] = []
    for tag, holders in skill_map.items():
        cnt = len(holders)
        if cnt == 1:
            risks.append(RiskWarningItem(
                level="red", type="skill",
                title=f"关键技能「{tag}」仅 1 人掌握",
                description="该技能仅 1 人掌握，存在严重单点风险",
                details=[f"{h['name']}（{h['ehr_no']}）" for h in holders],
            ))
        elif 2 <= cnt <= SKILL_YELLOW_MAX:
            risks.append(RiskWarningItem(
                level="yellow", type="skill",
                title=f"关键技能「{tag}」仅 {cnt} 人掌握",
                description="建议加强培训或招聘",
                details=[f"{h['name']}（{h['ehr_no']}）" for h in holders],
            ))
        else:
            risks.append(RiskWarningItem(
                level="green", type="skill",
                title=f"关键技能「{tag}」{cnt} 人掌握，能力充足",
                description="",
                details=[f"{h['name']}（{h['ehr_no']}）" for h in holders],
            ))
    # 排序：红 > 黄 > 绿
    order = {"red": 0, "yellow": 1, "green": 2}
    risks.sort(key=lambda r: (order.get(r.level, 9), r.type))
    return risks


async def _calc_emergency_risk(db: AsyncSession, current_user: User) -> RiskWarningItem:
    """应急能力风险（按整体检查）"""
    emg_q = (
        select(func.count(User.id))
        .join(Profile, Profile.user_id == User.id)
        .where(User.deleted_at.is_(None), Profile.is_emergency_staff == True)
    )
    emg_q = await _scope_filter(emg_q, db, current_user)
    cnt = (await db.execute(emg_q)).scalar() or 0

    total_q = select(func.count(User.id)).where(User.deleted_at.is_(None))
    total_q = await _scope_filter(total_q, db, current_user)
    total = (await db.execute(total_q)).scalar() or 0

    detail = f"当前 {cnt} 人 / 全员 {total} 人"
    if cnt < EMERGENCY_RED_MAX:
        return RiskWarningItem(
            level="red", type="emergency",
            title=f"应急先锋队仅 {cnt} 人",
            description="应急响应能力严重不足，建议立即补充",
            details=[detail],
        )
    elif cnt < EMERGENCY_YELLOW_MAX:
        return RiskWarningItem(
            level="yellow", type="emergency",
            title=f"应急先锋队 {cnt} 人",
            description="应急响应能力偏少，建议扩充",
            details=[detail],
        )
    else:
        return RiskWarningItem(
            level="green", type="emergency",
            title=f"应急先锋队 {cnt} 人，能力充足",
            description="",
            details=[detail],
        )


@router.post("/risks", response_model=RisksResponse)
async def risks(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ("admin", "leader"):
        raise HTTPException(status_code=403, detail="无权限")

    items = await _calc_skill_risks(db, current_user)
    items.append(await _calc_emergency_risk(db, current_user))
    red = sum(1 for r in items if r.level == "red")
    yellow = sum(1 for r in items if r.level == "yellow")
    green = sum(1 for r in items if r.level == "green")
    return RisksResponse(items=items, red_count=red, yellow_count=yellow, green_count=green)


# ======================== 4. emergency ========================

@router.post("/emergency", response_model=EmergencyStatsResponse)
async def emergency_stats(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ("admin", "leader"):
        raise HTTPException(status_code=403, detail="无权限")

    q = (
        select(User.ehr_no, ContactInfo.commute_minutes)
        .join(Profile, Profile.user_id == User.id)
        .outerjoin(ContactInfo, ContactInfo.profile_id == Profile.id)
        .where(User.deleted_at.is_(None), Profile.is_emergency_staff == True)
    )
    q = await _scope_filter(q, db, current_user)
    rows = (await db.execute(q)).all()

    bucket_map = {"<30分钟": 0, "30-60分钟": 0, ">60分钟": 0, "未设置": 0}
    for _, minutes in rows:
        if minutes is None:
            bucket_map["未设置"] += 1
        elif minutes < 30:
            bucket_map["<30分钟"] += 1
        elif minutes <= 60:
            bucket_map["30-60分钟"] += 1
        else:
            bucket_map[">60分钟"] += 1
    items = [EmergencyStatItem(bucket=b, count=c) for b, c in bucket_map.items()]

    total_q = select(func.count(User.id)).where(User.deleted_at.is_(None))
    total_q = await _scope_filter(total_q, db, current_user)
    total = (await db.execute(total_q)).scalar() or 0
    total_emg = sum(bucket_map.values())
    coverage = round((total_emg / total * 100) if total else 0.0, 1)

    return EmergencyStatsResponse(
        items=items,
        total_emergency=total_emg,
        total_employees=total,
        coverage_rate=coverage,
    )


# ======================== 5. export ========================

def _build_excel(
    overview: OverviewResponse,
    certs: CertificateStatResponse,
    risks: RisksResponse,
    emg: EmergencyStatsResponse,
) -> bytes:
    wb = Workbook()
    # Sheet 1: 能力分布
    ws1 = wb.active
    ws1.title = "能力分布"
    ws1.append(["技能名", "人数", "持有人 EHR"])
    for s in overview.skills:
        ws1.append([s.skill_name, s.count, ", ".join(s.holders)])
    # 组别密度
    ws1.append([])
    ws1.append(["组别", "总人数", "应急先锋队人数"])
    for g in overview.groups:
        ws1.append([g.group_name, g.count, g.is_emergency_count])

    # Sheet 2: 证书统计
    ws2 = wb.create_sheet("证书统计")
    ws2.append(["证书名", "人数", "持有人 EHR"])
    for c in certs.items:
        ws2.append([c.cert_name, c.count, ", ".join(c.holders)])

    # Sheet 3: 风险预警
    ws3 = wb.create_sheet("风险预警")
    ws3.append(["等级", "类型", "标题", "描述", "详情"])
    for r in risks.items:
        ws3.append([r.level, r.type, r.title, r.description, "\n".join(r.details)])

    # Sheet 4: 应急能力
    ws4 = wb.create_sheet("应急能力")
    ws4.append(["通勤区间", "人数"])
    for it in emg.items:
        ws4.append([it.bucket, it.count])
    ws4.append([])
    ws4.append(["应急先锋队总数", emg.total_emergency])
    ws4.append(["全员总数", emg.total_employees])
    ws4.append(["覆盖率（%）", emg.coverage_rate])

    # 表头样式
    for ws in [ws1, ws2, ws3, ws4]:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="D05A6E", end_color="D05A6E", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        # 列宽
        for col_letter, width in zip("ABCDE", [25, 15, 30, 50, 40]):
            ws.column_dimensions[col_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@router.post("/export")
async def export(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ("admin", "leader"):
        raise HTTPException(status_code=403, detail="无权限")

    ov = await overview(db, current_user)
    cr = await certificates(db, current_user)
    rk = await risks(db, current_user)
    em = await emergency_stats(db, current_user)
    data = _build_excel(ov, cr, rk, em)
    import urllib.parse
    fname_zh = f"团队能力分析报告_{date.today().strftime('%Y-%m-%d')}.xlsx"
    fname_ascii = f"analytics_{date.today().strftime('%Y%m%d')}.xlsx"
    fname_encoded = urllib.parse.quote(fname_zh)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{fname_ascii}"; '
                f"filename*=UTF-8''{fname_encoded}"
            ),
        },
    )
