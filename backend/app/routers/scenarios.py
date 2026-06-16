# Step 3: 智能筛选四大场景 API
# 包含：search 统一查询 + export Excel 导出
from __future__ import annotations

import io
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, exists
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models import (
    User,
    Profile,
    ContactInfo,
    ProfileSkillTag,
    QualificationInfo,
    ProjectSummary,
)
from app.schemas import (
    ScenarioSearchRequest,
    ScenarioSearchItem,
    ScenarioSearchResponse,
)
from app.auth import get_current_admin, leader_effective_group
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

router = APIRouter(prefix="/api/admin/scenarios", tags=["智能筛选场景"])


# ======================== 工具函数 ========================

def _clamp_score(v: float) -> float:
    try:
        return float(max(0.0, min(100.0, v)))
    except Exception:
        return 0.0


async def _list_users_for_scenario(
    db: AsyncSession,
    req: ScenarioSearchRequest,
    current_user: User,
):
    """基础查询：返回带 profile/contact/skill_tags/cert_count/project_count 的 User 列表（不分页）"""
    base_q = (
        select(User)
        .where(User.deleted_at.is_(None))
        .options(
            selectinload(User.profile).selectinload(Profile.contact),
            selectinload(User.profile).selectinload(Profile.skill_tags),
        )
    )
    if not req.include_disabled:
        base_q = base_q.where(User.is_disabled.is_(False))
    if req.group_name:
        base_q = base_q.where(User.group_name == req.group_name)
    if req.role:
        base_q = base_q.where(User.role == req.role)
    if current_user.role == "leader":
        lg = leader_effective_group(current_user)
        if not lg:
            raise HTTPException(status_code=403, detail="组长未配置有效组别")
        base_q = base_q.where(User.group_name == lg)
    result = await db.execute(base_q.order_by(User.id))
    return result.scalars().unique().all()


async def _attach_counts(db: AsyncSession, users: list[User]) -> dict[int, tuple[int, int]]:
    """批量获取所有用户的 cert_count 和 project_count"""
    user_ids = [u.id for u in users]
    if not user_ids:
        return {}
    cert_q = select(QualificationInfo.profile_id, func.count(QualificationInfo.id)).join(
        Profile, QualificationInfo.profile_id == Profile.id
    ).where(Profile.user_id.in_(user_ids)).group_by(QualificationInfo.profile_id)
    proj_q = select(ProjectSummary.profile_id, func.count(ProjectSummary.id)).join(
        Profile, ProjectSummary.profile_id == Profile.id
    ).where(Profile.user_id.in_(user_ids)).group_by(ProjectSummary.profile_id)
    cert_map = {pid: cnt for pid, cnt in (await db.execute(cert_q)).all()}
    proj_map = {pid: cnt for pid, cnt in (await db.execute(proj_q)).all()}
    profile_id_to_user_id = {}
    pids = [u.id for u in users]
    if pids:
        pid_rows = (await db.execute(select(Profile.id, Profile.user_id).where(Profile.user_id.in_(pids)))).all()
        for pid, uid in pid_rows:
            profile_id_to_user_id[pid] = uid
    out: dict[int, tuple[int, int]] = {}
    for u in users:
        profile = u.profile
        if not profile:
            out[u.id] = (0, 0)
            continue
        out[u.id] = (cert_map.get(profile.id, 0), proj_map.get(profile.id, 0))
    return out


# ======================== 场景查询 ========================

@router.post("/search", response_model=ScenarioSearchResponse)
async def scenario_search(
    body: ScenarioSearchRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_admin),
):
    if body.scenario not in ("emergency", "activity", "project", "transfer"):
        raise HTTPException(status_code=422, detail="未知场景类型")

    users = await _list_users_for_scenario(db, body, current_user)
    counts_map = await _attach_counts(db, users)

    items: list[ScenarioSearchItem] = []
    for u in users:
        profile = u.profile
        commute_minutes = profile.contact.commute_minutes if (profile and profile.contact) else None
        skill_tags = [t.tag_name for t in (profile.skill_tags if profile else [])]
        is_emergency = bool(profile and profile.is_emergency_staff)
        cert_count, project_count = counts_map.get(u.id, (0, 0))

        item = ScenarioSearchItem(
            ehr_no=u.ehr_no,
            name=u.name,
            group_name=u.group_name,
            role=u.role,
            is_emergency_staff=is_emergency,
            commute_minutes=commute_minutes,
            skill_tags=skill_tags,
            cert_count=cert_count,
            project_count=project_count,
        )

        if body.scenario == "emergency":
            # Q1: 应急置底 = 不隐藏，按 is_emergency_staff 降序 + 通勤升序
            if body.max_commute_minutes is not None:
                if commute_minutes is None or commute_minutes >= body.max_commute_minutes:
                    continue
            item.match_score = _clamp_score(100 if is_emergency else 0)
        elif body.scenario == "activity":
            if not body.interest_tags:
                raise HTTPException(status_code=422, detail="活动选人需指定 interest_tags")
            matched = [t for t in body.interest_tags if t in skill_tags]
            if not matched:
                continue
            item.matched_tags = matched
            item.has_interests = True
            # 评分：每个命中 20 分，封顶 100
            item.match_score = _clamp_score(len(matched) * 20)
        elif body.scenario == "project":
            req_skills = body.required_skill_tags or []
            min_cert = body.min_cert_count or 0
            min_proj = body.min_project_count or 0
            score = 0
            matched_skills = [s for s in req_skills if s in skill_tags]
            if req_skills and not matched_skills:
                continue  # 必须有必选技能
            score += len(matched_skills) * 30
            if cert_count >= min_cert:
                score += 20
            if project_count >= min_proj:
                score += 20
            if commute_minutes is not None and commute_minutes < 60:
                score += 10
            item.match_score = _clamp_score(score)
            item.matched_tags = matched_skills
            item.has_required_skills = bool(matched_skills)
        elif body.scenario == "transfer":
            if not body.target_group:
                raise HTTPException(status_code=422, detail="人员调配需指定 target_group")
            if u.group_name == body.target_group:
                continue  # 不能调配到原组
            if body.max_commute_minutes is not None:
                if commute_minutes is None or commute_minutes > body.max_commute_minutes:
                    continue
            # 评分：技能数 + 应急 + 通勤
            score = 0
            score += min(len(skill_tags) * 5, 40)
            if is_emergency:
                score += 20
            if commute_minutes is not None:
                if commute_minutes < 30:
                    score += 30
                elif commute_minutes < 60:
                    score += 15
            item.match_score = _clamp_score(score)

        items.append(item)

    # 排序
    if body.scenario == "emergency":
        items.sort(key=lambda x: (not x.is_emergency_staff, x.commute_minutes if x.commute_minutes is not None else 999999, x.ehr_no))
    else:
        items.sort(key=lambda x: (-x.match_score, x.commute_minutes if x.commute_minutes is not None else 999999, x.ehr_no))

    return ScenarioSearchResponse(total=len(items), items=items, scenario=body.scenario)


# ======================== 导出 Excel ========================

SCENARIO_FILE_NAMES = {
    "emergency": "应急人员清单",
    "activity": "活动选人",
    "project": "项目组队",
    "transfer": "人员调配",
}


def _build_excel(items: list[ScenarioSearchItem], scenario: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = SCENARIO_FILE_NAMES.get(scenario, "场景结果")

    headers = [
        "EHR 号", "姓名", "组别", "角色",
        "应急先锋队", "通勤时间(分钟)", "技能标签",
        "证书数", "项目数", "匹配度",
    ]
    ws.append(headers)
    # 表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="D05A6E", end_color="D05A6E", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for it in items:
        ws.append([
            it.ehr_no,
            it.name,
            it.group_name,
            it.role,
            "是" if it.is_emergency_staff else "否",
            it.commute_minutes if it.commute_minutes is not None else "",
            ", ".join(it.skill_tags),
            it.cert_count,
            it.project_count,
            round(it.match_score, 1),
        ])

    # 列宽
    widths = [12, 12, 18, 12, 12, 14, 40, 10, 10, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@router.post("/export")
async def scenario_export(
    body: ScenarioSearchRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_admin),
):
    """复用 search 逻辑，结果导出 xlsx"""
    # 复用 search 内部逻辑
    res = await scenario_search(body, db, current_user)
    data = _build_excel(res.items, body.scenario)
    import urllib.parse
    fname_zh = f"{SCENARIO_FILE_NAMES.get(body.scenario, '场景')}_{date.today().strftime('%Y-%m-%d')}.xlsx"
    fname_ascii = f"scenario_{body.scenario}_{date.today().strftime('%Y%m%d')}.xlsx"
    # RFC 5987 编码中文文件名（兼容性更好）
    fname_encoded = urllib.parse.quote(fname_zh)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{fname_ascii}"; '
                f"filename*=UTF-8''{fname_encoded}"
            ),
        },
    )
