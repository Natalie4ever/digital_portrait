# 用户管理：仅管理员。增删改查、批量导入、重置密码、启用/禁用
from __future__ import annotations

from typing import Optional
import io
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, exists
from sqlalchemy.orm import selectinload
from openpyxl import load_workbook

from app.database import get_db
from app.models import User, Profile
from app.schemas import (
    UserCreate,
    UserUpdate,
    UserResponse,
    UserListResponse,
    AdminResetPasswordRequest,
)
from app.auth import get_current_admin, get_current_user, leader_effective_group, hash_password, validate_password_strength
from app.config import settings
from app.operation_log import log_operation
from app.validators import validate_ehr_no

router = APIRouter(prefix="/api/admin/users", tags=["用户管理"])


@router.get("", response_model=UserListResponse)
async def list_users(
    page: int = 1,
    page_size: int = 20,
    ehr_no: Optional[str] = None,
    name: Optional[str] = None,
    group_name: Optional[str] = None,
    role: Optional[str] = None,
    include_disabled: bool = True,
    is_emergency_staff: Optional[bool] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_admin),
):
    q = select(User).where(User.deleted_at.is_(None)).options(selectinload(User.profile))
    if not include_disabled:
        q = q.where(User.is_disabled == False)
    if ehr_no:
        q = q.where(User.ehr_no.contains(ehr_no))
    if name:
        q = q.where(User.name.contains(name))
    if group_name:
        q = q.where(User.group_name == group_name)
    if role:
        q = q.where(User.role == role)
    if is_emergency_staff is not None:
        q = q.where(
            exists().where(Profile.user_id == User.id, Profile.is_emergency_staff == is_emergency_staff)
        )
    count_q = select(func.count()).select_from(User).where(User.deleted_at.is_(None))
    if not include_disabled:
        count_q = count_q.where(User.is_disabled == False)
    if ehr_no:
        count_q = count_q.where(User.ehr_no.contains(ehr_no))
    if name:
        count_q = count_q.where(User.name.contains(name))
    if group_name:
        count_q = count_q.where(User.group_name == group_name)
    if role:
        count_q = count_q.where(User.role == role)
    if is_emergency_staff is not None:
        count_q = count_q.where(
            exists().where(Profile.user_id == User.id, Profile.is_emergency_staff == is_emergency_staff)
        )
    total = (await db.execute(count_q)).scalar() or 0
    q = q.offset((page - 1) * page_size).limit(page_size).order_by(User.id)
    result = await db.execute(q)
    items = result.scalars().all()
    return UserListResponse(
        total=total,
        items=[UserResponse.model_validate(u) for u in items],
    )


@router.post("", response_model=UserResponse)
async def create_user(
    body: UserCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_admin),
):
    r = await db.execute(select(User).where(User.ehr_no == body.ehr_no.strip()))
    if r.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="EHR号已存在")
    pwd = body.initial_password or settings.DEFAULT_PASSWORD
    user = User(
        ehr_no=body.ehr_no.strip(),
        name=body.name.strip(),
        group_name=body.group_name.strip(),
        role=body.role if body.role in ("user", "leader", "admin") else "user",
        password_hash=hash_password(pwd),
    )
    db.add(user)
    await db.flush()
    await log_operation(db, current_user.id, "create_user", "users", user.ehr_no, None)
    return UserResponse.model_validate(user)


def _parse_ehr_path(ehr_no: str) -> str:
    try:
        return validate_ehr_no(ehr_no)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))


@router.get("/{ehr_no}", response_model=UserResponse)
async def get_user(
    ehr_no: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_admin),
):
    ehr_no = _parse_ehr_path(ehr_no)
    r = await db.execute(select(User).where(User.ehr_no == ehr_no, User.deleted_at.is_(None)))
    user = r.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    return UserResponse.model_validate(user)


@router.put("/{ehr_no}", response_model=UserResponse)
async def update_user(
    ehr_no: str,
    body: UserUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_admin),
):
    ehr_no = _parse_ehr_path(ehr_no)
    r = await db.execute(select(User).where(User.ehr_no == ehr_no, User.deleted_at.is_(None)))
    user = r.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    if body.name is not None:
        user.name = body.name.strip()
    if body.role is not None and body.role in ("user", "leader", "admin"):
        user.role = body.role
    if body.is_disabled is not None:
        user.is_disabled = body.is_disabled
    # Step 2: 调组（group_name 变化）走 transfer_user 写历史
    if body.group_name is not None and body.group_name.strip() != (user.group_name or "").strip():
        from app.routers.group_transfers import _do_transfer
        await _do_transfer(
            db,
            ehr_no=ehr_no,
            to_group=body.group_name,
            transfer_date=None,
            reason="通过用户管理编辑调组",
            remark=None,
            operator=current_user,
        )
    elif body.group_name is not None:
        # 同组内仅做 trim，不写历史
        user.group_name = body.group_name.strip()
    db.add(user)
    await db.flush()
    await log_operation(db, current_user.id, "update_user", "users", ehr_no, None)
    return UserResponse.model_validate(user)


@router.delete("/{ehr_no}")
async def delete_user(
    ehr_no: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_admin),
):
    from datetime import datetime
    ehr_no = _parse_ehr_path(ehr_no)
    r = await db.execute(select(User).where(User.ehr_no == ehr_no, User.deleted_at.is_(None)))
    user = r.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    user.deleted_at = datetime.utcnow()
    db.add(user)
    await db.flush()
    await log_operation(db, current_user.id, "delete_user", "users", ehr_no, None)
    return {"message": "已删除（软删除）"}


@router.post("/reset-password")
async def reset_password(
    body: AdminResetPasswordRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_admin),
):
    r = await db.execute(select(User).where(User.ehr_no == body.ehr_no, User.deleted_at.is_(None)))
    user = r.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    user.password_hash = hash_password(body.new_password)
    db.add(user)
    await db.flush()
    await log_operation(db, current_user.id, "reset_password", "users", body.ehr_no, None)
    return {"message": "密码已重置"}


@router.post("/batch-import")
async def batch_import(
    file: UploadFile,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_admin),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 Excel 文件（.xlsx）")
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        raise HTTPException(status_code=400, detail="Excel 无有效工作表")
    # 表头：姓名、ehr号、组别、角色（可选）、初始密码（可选）
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel 为空")
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    name_col = ehr_col = group_col = role_col = pwd_col = -1
    for i, h in enumerate(header):
        if "姓名" in h or h == "name":
            name_col = i
        elif "ehr" in h.lower() or "工号" in h:
            ehr_col = i
        elif "组别" in h or "组" in h:
            group_col = i
        elif "角色" in h or "role" in h.lower():
            role_col = i
        elif "密码" in h or "初始密码" in h:
            pwd_col = i
    if name_col < 0 or ehr_col < 0 or group_col < 0:
        raise HTTPException(status_code=400, detail="Excel 需包含列：姓名、ehr号、组别")
    created = 0
    skipped = 0
    errors = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if not row:
            continue
        name = str(row[name_col]).strip() if name_col < len(row) and row[name_col] else ""
        ehr = str(row[ehr_col]).strip() if ehr_col < len(row) and row[ehr_col] else ""
        group = str(row[group_col]).strip() if group_col < len(row) and row[group_col] else ""
        role = str(row[role_col]).strip() if role_col >= 0 and role_col < len(row) and row[role_col] else "user"
        pwd = str(row[pwd_col]).strip() if pwd_col >= 0 and pwd_col < len(row) and row[pwd_col] else ""
        if not ehr:
            continue
        try:
            ehr = validate_ehr_no(ehr)
        except ValueError:
            errors.append(f"第{row_idx}行: EHR号必须为7位数字")
            continue
        if role not in ("user", "leader", "admin"):
            role = "user"
        r = await db.execute(select(User).where(User.ehr_no == ehr))
        if r.scalar_one_or_none():
            skipped += 1
            continue
        if pwd:
            ok, msg = validate_password_strength(pwd)
            if not ok:
                errors.append(f"第{row_idx}行: 密码强度不足 - {msg}")
                continue
        else:
            pwd = settings.DEFAULT_PASSWORD
        user = User(
            ehr_no=ehr,
            name=name or ehr,
            group_name=group or "未分组",
            role=role,
            password_hash=hash_password(pwd),
        )
        db.add(user)
        created += 1
    await db.flush()
    await log_operation(db, current_user.id, "batch_import", "users", f"导入: 新增 {created}, 跳过 {skipped}", None)
    return {"message": "导入完成", "created": created, "skipped": skipped, "errors": errors}


# ===================== Step 1 1.1（修订）: 切换应急先锋队标识 =====================
async def _toggle_emergency(db: AsyncSession, ehr_no: str, current_user: User, log_action: str):
    """统一的应急先锋队切换逻辑：admin 可切任意人，leader 仅可切本组"""
    ehr_no = _parse_ehr_path(ehr_no)
    r = await db.execute(
        select(User)
        .where(User.ehr_no == ehr_no, User.deleted_at.is_(None))
        .options(selectinload(User.profile))
    )
    user = r.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    # 组长只能标记本组
    if current_user.role == "leader":
        lg = leader_effective_group(current_user)
        if not lg or (user.group_name or "").strip() != lg:
            raise HTTPException(status_code=403, detail="仅可标记本组成员")
    # 普通员工/组长调用 admin-only 接口会被 get_current_admin 拦掉（仅 admin 可访问）
    profile = user.profile
    if not profile:
        profile = Profile(user_id=user.id)
        db.add(profile)
        await db.flush()
    profile.is_emergency_staff = not bool(profile.is_emergency_staff)
    await log_operation(
        db, current_user.id, log_action, "profile",
        f"{ehr_no} -> is_emergency_staff={profile.is_emergency_staff}", None,
    )
    return {"ehr_no": ehr_no, "is_emergency_staff": profile.is_emergency_staff}


@router.post("/{ehr_no}/toggle-emergency")
async def toggle_emergency_user(
    ehr_no: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_admin),
):
    """用户管理列表调用：仅 admin 可调用"""
    return await _toggle_emergency(db, ehr_no, current_user, "toggle_emergency_staff")
